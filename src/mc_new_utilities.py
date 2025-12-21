#!/usr/bin/env python
""" Common utilities for MonteCarlo  """

# import re
# import sys
# import os

import pandas
import pandas as pd
# import datetime as dt
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import string

DEBUG = False
Money_format = '$#,##0.00_);[Red]($#,##0.00)'  # Excel cell number format for dollar values
Zoom_level = 110
XL_col_width = 11  # Width of a column for $ in Excel
XL_col_1_width = 18  # Width of 1st column


def linear_transform_fastest(M_in, slope, intercept):
    M = M_in.copy(deep=True)  # avoid modifying the input matrix
    for i in M.index:
        M.loc[i, :] *= slope[i]
        M.loc[i, :] += intercept[i]
    return M


def float_2_pct(in_nb: float, nb_decimal: int = 2) -> str:
    """ Returns a string converting a float number to pct - e.g. 0.0241 -> '2.41%' """
    return f"{100.0 * in_nb:,.{nb_decimal}f}%"


def mv_last_n_to_front(lst, n):
    """ Moves the last n elements of a list to first positions - in the same order """
    return lst[len(lst) - n:] + lst[:-n]


def mv_last_n_col_to_front(df, n):
    """ Moves the last n columns of a DF to first positions - in the same order """
    col = list(df.columns)
    return df[mv_last_n_to_front(col, n)]


def expand_dimensions(dim_str):
    """
    :param dim_str: format 'A1:XYddd' where ddd is a number and XY a set of letters
    :return:
    """
    t_l, b_r = dim_str.split(':')
    assert t_l == "A1", f"Whaaat??? top left {t_l} is not A1"
    # figure out the label bottom right cell
    last_col = ''
    digits = []
    for c in b_r:
        if c.isalpha():
            last_col += c
        elif c.isnumeric():
            digits += [c]
        else:
            print(f"Wow we have an unexpected character: {c}")
    # print(f"digits: {digits}")
    # Convert the list of digits into an actual number
    digits.reverse()  # list digits in increasing power of 10
    powr_10 = 1
    last_row = 0
    for d in digits:
        last_row += int(d) * powr_10
        powr_10 *= 10

    # Make a list of all the letter labels of the columns
    assert len(last_col) <= 2, f"FixMe: I can only deal with 1-2 letter column labels not {len(last_col)} - last_col " \
                               f"= {last_col}"
    col_range = []
    # Start with single letter labels
    for ll in string.ascii_uppercase:
        col_range += [ll]
        if ll == last_col:
            break
    if len(last_col) == 2:  # 2-letter column labels - if needed
        for l1 in string.ascii_uppercase:
            for l2 in string.ascii_uppercase:
                ll = l1 + l2
                col_range += [ll]
                if ll == last_col:
                    break
    return col_range, last_row


def write_dollar_df_2_xl(xl_file: str, df: pandas.DataFrame, sheet: str, a_w_mode: str = 'a', index: bool = False,
                         mv_last: bool = None) -> None:
    """
    Writes to an Excel worksheet and formats nicely a DF whose values represent $
    FYI Need both writer and file name
    """
    if mv_last:  # Move last columns to front
        df = mv_last_n_col_to_front(df, mv_last)
    with pd.ExcelWriter(xl_file, mode=a_w_mode, engine="openpyxl") as xlw:
        df.to_excel(xlw, sheet_name=sheet, float_format='%.2f', header=True, index=index)

    # Format the workbook
    workbook = openpyxl.load_workbook(xl_file)
    sheet_idx = workbook.sheetnames.index(sheet)
    ws = workbook.worksheets[sheet_idx]
    dim_str = ws.dimensions
    col_names, max_row = expand_dimensions(dim_str)
    left_top_align = Alignment(horizontal='left', vertical='top')
    center_bottom_align = Alignment(horizontal='center', vertical='bottom')
    # Make 1st col bold and left-aligned
    for row in range(1, max_row + 1):
        cel = 'A' + str(row)
        ws[cel].font = Font(bold=True)
        ws[cel].alignment = left_top_align
    # Make 1st row bold and center-aligned
    for col in col_names:
        cel = col + str(1)
        ws[cel].font = Font(bold=True)
        ws[cel].alignment = center_bottom_align
    for col in ws.iter_cols(min_row=2, max_row=max_row, min_col=2, max_col=len(col_names)):
        for cel in col:
            cel.number_format = '_("$"#,##0.00_);[Red]("$"#,##0.00);_($* "-"??_);_(@_)'

    # Set column width
    for col in col_names:
        # Reference: https://www.codespeedy.com/change-or-modify-column-width-size-in-openpyxl/
        # Reference: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.dimensions.html
        ws.column_dimensions[col].bestFit = True

    # Freeze header row and column
    ws.freeze_panes = ws['B2']
    ws.sheet_view.zoomScale = Zoom_level
    workbook.save(xl_file)
    return


def print_df(df, msg=None, verbose=False):
    """ Print troubleshooting info on a DF """
    msg_to_print = msg if msg is not None else ""
    print(f"\n*** {msg_to_print}: {len(df.index)} rows")
    if type(df) == pandas.core.frame.DataFrame:
        print(df.head())
        # print(f"Columns: {df.columns}")
        if verbose:
            print(f"Columns: {', '.join(df.columns)}")
            print(f"Index:  {', '.join(df.index)}")
            print('/* ... */')
            print(df.tail())
            print('/* --------- */')
    elif type(df) == pandas.core.frame.Series:
        # print(f"{df.name}:")
        # print(*df.items(), sep='\n')
        print(f"{df.name}: {dict(df)}")
    else:
        print(f"{msg_to_print}: Array neither DF nor Series - skipping print")
    return


def print_ser_pct(in_ser: pd.Series) -> dict:
    """
    Returns a dictionary where key is index of series, and value is value of item converted to a % string
    0.2512 -> "25.12%"
    @param in_ser: input series
    @return: dict with nicely formatted values
    """
    return {ky: f"{100 * val:,.2f}%" for ky, val in in_ser.items()}


def print_ser_float(in_ser: pd.Series, nb_decimal: int = 0) -> dict:
    """
    Returns a dictionary where key is index of series, and value is string with nicely formatted float
    1111.1111 -> "1,111.11"
    @param in_ser: input series
    @param nb_decimal: number of decimals in the printout
    @return: dict with nicely formatted values
    """
    assert nb_decimal >= 0, f"print_ser_float number decimals ({nb_decimal}) must be >= 0"
    return {ky: f"{val:,.{nb_decimal}f}" for ky, val in in_ser.items()}


def print_df_dollar(in_df: pd.DataFrame, nb_decimal=0) -> pd.DataFrame:
    """
    Prints a DF that contains $ amounts nicely
    @param in_df: input DF
    @param nb_decimal: # decimals to display
    @return: DF containing strings
    """
    out_df = in_df.copy(deep=True)
    return out_df.applymap(lambda x: f"${x:,.{nb_decimal}f}")


def print_out(outfile, msg='\n'):
    """ Prints message to both outfile and terminal
    :rtype: None
    """
    print(msg)
    outfile.write(msg)
    return
