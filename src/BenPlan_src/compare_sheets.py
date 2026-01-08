#!/usr/bin/env python3
""""
Compare the Excel files from BenPlan and benplan_oo
Note that the comparison is not totally correct: the order of the columns is not checked.
"""

import pandas as pd
import sys
import datetime as dt
import os
import re
from openpyxl import load_workbook
from pandas.api.types import is_string_dtype


HOME_DIR = '/Users/bfraenkel/Documents/Code/BenPlan/'
DATA_DIR = HOME_DIR + '/Data/'
DIR_NAME_PATTERN = '\d{4}-\d{2}-\d{2}'

latest_dir = ""  # place holder for the latest directory

def diff_list(list1: list, list2: list, title: str) -> list:
    """
    Compute the difference between the two lists.
    Returns True if the lists are the same, False otherwise.
    Note that this function does not check for the order of the elements in the lists.
    """
    # show the elements in both lists, those in list1 but not in list2 and those in list2 but not in list1
    l1_only = [x for x in list1 if x not in list2]
    l2_only = [x for x in list2 if x not in list1]
    both = [x for x in list1 if x in list2]
    if len(both) == len(list1):
        return True
    elif len(both) == 0:
        print(f"{title} - Lists have NOTHING in common")
        return False
    else:
        if len(l1_only) > 0:
            print(f"{title} - l1_only: {l1_only}")
        if len(l2_only) > 0:
            print(f"{title} - l2_only: {l2_only}")
        # if len(both) > 0:
        #     print(f"both: {both}")
        return False

def diff_df(df1: pd.DataFrame, df2: pd.DataFrame, title: str) -> (pd.DataFrame, bool):
    """Compute the difference between the two dataframes.
    Knowing that they have the same columns and indexes,and the values are numeric,
    """
    # First, make sure the indexes are the same and the columns are the same
    index_flag = diff_list(df1.index, df2.index, title)
    columns_flag = diff_list(df1.columns, df2.columns, title)
    if not index_flag or not columns_flag:
        print(f"{title} - Indexes or columns are not the same")
        return None, False
    diff_df = pd.DataFrame(index=df1.index, columns=df1.columns)
    error_cols = []
    for col in df1.columns:
        try:
            diff_df[col] = df1[col] - df2[col]
            if diff_df[col].sum() != 0:
                error_cols.append(col)
        except Exception as e: # print the exception
            if is_string_dtype(df1[col]) and is_string_dtype(df2[col]):
                # compare the strings in the two columns    
                if list(df1[col]) == list(df2[col]):
                    diff_df[col] = 0        
                else:
                    print(f"{title} - String column {col} is different")
                    diff_df[col] = df1[col] 
                    error_cols.append(col)
            else:
                print(f"{title} - Error in column {col}: {e}")  # print the exception
                diff_df[col] = df1[col] 

    if len(error_cols) > 0:
        print(f"{title} - Error columns: {error_cols}")
    else:
        print(f"{title} - Different but no delta (colomns out of order?)")
    return diff_df, True



def get_current_directory() -> str:
    """Get the most recent data directory."""
    d_files = []
    with os.scandir(DATA_DIR) as dir_list:
        for file_or_dir in dir_list:
            if file_or_dir.is_dir():
                d_files.append(file_or_dir.name)

    dir_files = [d for d in d_files if re.fullmatch(DIR_NAME_PATTERN, d)]
    return max(dir_files)

def same_sheet_names(bp_file: str, bp_oo_file: str) -> bool:
    """Check if the sheet names are the same in the two Excel files."""
    # get the sheet names from the two files
    wb = load_workbook(bp_file, read_only=True)
    bp_sheet_names = wb.sheetnames
    wb_oo = load_workbook(bp_oo_file, read_only=True)
    bp_oo_sheet_names = wb_oo.sheetnames
    # find sheets that are in bp_file but not in bp_oo_file
    missing_sheets = [s for s in bp_sheet_names if s not in bp_oo_sheet_names]
    # find sheets that are in bp_oo_file but not in bp_file
    extra_sheets = [s for s in bp_oo_sheet_names if s not in bp_sheet_names]
    common_sheets = [s for s in bp_sheet_names if s in bp_oo_sheet_names]
    return missing_sheets, extra_sheets, common_sheets

def compare_sheets(bp_file: str, bp_oo_file: str, sheet_name: str) -> bool:
    """Compare the two sheets with the same name."""
    # 1st row is the header, 1st column is the index
    df_bp = pd.read_excel(bp_file, sheet_name=sheet_name, header=0, index_col=0)
    df_oo = pd.read_excel(bp_oo_file, sheet_name=sheet_name, header=0, index_col=0)
    # compare the two dataframes
    if df_bp.equals(df_oo):
        return True
    else:
        # Compute the difference between the two dataframes and save it to a new file
        delta_df, equal_flag = diff_df(df_bp, df_oo, sheet_name)
        if equal_flag:
            return True
        elif delta_df is None:
            print(f"Error in sheet: {sheet_name}")
            return False
        else:
            # Save the difference dataframe to a new file
            delta_df.to_excel(f'{DATA_DIR}/{latest_dir}/{sheet_name}_delta.xlsx')
            return False
        

def main(cmd_line: list[str]):
    # find the latest directory
    global latest_dir

    latest_dir = get_current_directory()
    print(f"Latest directory: {latest_dir}")

    # get the Excel files from the latest directory
    bp_file = DATA_DIR + latest_dir + '/BenPlan-' + latest_dir + '.xlsx'
    bp_oo_file = DATA_DIR + latest_dir + '/benplan_oo-' + latest_dir + '.xlsx'
    missing_sheets, extra_sheets, common_sheets = same_sheet_names(bp_file, bp_oo_file)
    if missing_sheets:
        print(f"Missing sheets: {missing_sheets}")
    else:
        print("No missing sheets")
    if extra_sheets:
        print(f"Extra sheets: {extra_sheets}")
    else:
        print("No extra sheets")
    if  common_sheets:
        print(f"Common sheets: {common_sheets}")
    else:
        print("No common sheets")

    # Compare the data in the sheets that are in both files
    diff_sheets = []
    same_sheets = []
    for sheet_name in common_sheets:
        if not compare_sheets(bp_file, bp_oo_file, sheet_name):
            diff_sheets.append(sheet_name)
        else:
            same_sheets.append(sheet_name)
    if diff_sheets:
        print(f"Different sheets:")
        for sheet_name in diff_sheets:
            print(f"    {sheet_name}")
    else:
        print("No sheets are different")
    if same_sheets:
        print(f"Same sheets:")
        for sheet_name in same_sheets:
            print(f"    {sheet_name}")
    else:
        print("No same sheets")
    return

if __name__ == "__main__":
    # execute only if run as a script
    start_time = dt.datetime.now()
    main(sys.argv[1:])
    end_time = dt.datetime.now()
    print(f'End: {str(end_time)} -- Run Time: {str(end_time - start_time)}\n')
    sys.exit(0)
