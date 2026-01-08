#!/Library/Frameworks/Python.framework/Versions/Current/bin/python3

"""
Aggregate Venmo Statement files with Quicken files and others
"""
import os
import re
import string
import sys

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font

BFName = 'Bernard Fraenkel'

ZOOM_LEVEL = 110  # Zoom level for the Venmo files
# columns that we use in the worksheet
col_wrksheet = ["Date", "Category", "Payee", "Amount", "Tags", "Memo/Notes"]
# columns that we use in the Venmo file
VenmoCol2Keep = ["Datetime", "Type", "Note", "From", "To", "Amount (total)"]
VENMO_HEADER = 2
VENMO_NAME_PATTERN = "Venmo_[0-9]{4}_[0-9]{2}\\.csv"
VENMO_STATEMENT_SHEET = "Statements"
MapDict = {"Datetime": "Date", "Payee": "Payee", "Note": "Memo/Notes", "Amount": "Amount", }
BenPlanCol = ["Date", "Payee", "Category", "Tags", "Memo/Notes", "Amount", "Source"]


def set_payee(ser3: pd.Series) -> str:
    """ Determine if it's a transfer back to Chase or a payment """
    if ser3["Type"] == "Standard Transfer":  # Transfer $$ back to Chase account
        return "To Chase"
    # else this is a payment to or from me
    list_2 = list(ser3[1:])  # drop 'Type'
    assert BFName in list_2, f"{BFName} is neither payee nor payor: {list_2}"
    return list_2[1] if list_2[0] == BFName else list_2[0]


def set_amount(amount_text: str) -> float:
    """
    Convert string of format '-$1,234.56' to a float -1234.56
    or '$1,234.56' | '+$1,234.56'  to 1234.56
    """
    sign, value = amount_text.split("$")
    sign = sign.strip()  # remove blanks
    value = re.sub(",", "", value)  # remove commas
    assert sign in ["+", "-", ""], f"Unrecognized Amount: {amount_text}"
    amount = -float(value) if sign == "-" else float(value)
    return amount


def is_venmo_file_recent(venmo_file_name: str, yyyy_mm: str) -> bool:
    """
    Determine if a Venmo file with name syntax "Venmo_yyyy_mm.csv" is more recent than
    the yyyy_mm string argument
    :param venmo_file_name: Venmo file name
    :param yyyy_mm: Month to check against
    :return: True/False - based on whether the file is more recent than the yyyy_mm string
    """
    file_month = re.sub("Venmo_", "", re.sub(".csv", "", venmo_file_name))
    return file_month > yyyy_mm  # True if the file is more recent e.g. '2022_09' > '2022_08'


def make_list_of_files_to_process(file_lst: [str], yyyy_mm: str) -> [str]:
    """
    :param file_lst: list of Venmo files with name syntax "Venmo_yyyy_mm.csv"
    :param yyyy_mm: Month string
    :return: list of files that are more recent than the argument yyyy_mm aka Year_Month
    """
    return [f for f in file_lst if is_venmo_file_recent(f, yyyy_mm)]


def show_values(labels: [str], values: pd.Series) -> str:
    """
    Display labels and values of a series in a 1-line string
    :param labels:
    :param values:
    :return:
    """
    out_str = ""
    for l, v in zip(labels, list(values)):
        out_str += f"{l}: {v} - "
    return out_str


def expand_dimensions(dim_str):
    """
    :param dim_str: format 'A1:XYddd' where ddd is a number and XY a set of letters
    :return: [A, B, ..., XY], ddd where
    [A, B, ..., XY] is the list of letters (or letter combos) from 'A' to 'XY and
    ddd is the last row
    """
    t_l, b_r = dim_str.split(":")
    assert t_l == "A1", f"Whaaat??? top left {t_l} is not A1"
    # figure out the label bottom right cell
    last_col = ""
    digits = []
    for c in b_r:
        if c.isalpha():
            last_col += c
        elif c.isnumeric():
            digits += [c]
        else:
            print(f"Wow we have an unexpected character: {c}")
    # print(f"digits: {digits}")
    # Convert the list of digits into an actual number
    digits.reverse()  # list digits in increasing power of 10
    powr_10 = 1
    last_row = 0
    for d in digits:
        last_row += int(d) * powr_10
        powr_10 *= 10

    # Make a list of all the letter labels of the columns
    assert len(last_col) <= 2, f"FixMe: I can only deal with 1-2 letter column labels not  \
            {len(last_col)} - last_col = {last_col}"
    col_range = []
    # Start with single letter labels
    for ll in string.ascii_uppercase:
        col_range += [ll]
        if ll == last_col:
            break
    if len(last_col) == 2:  # 2-letter column labels - if needed
        for l1 in string.ascii_uppercase:
            for l2 in string.ascii_uppercase:
                ll = l1 + l2
                col_range += [ll]
                if ll == last_col:
                    break
    return col_range, last_row


def map_usdate_to_sortable_date(us_date: str) -> str:
    """Convert mm/dd/yyyy to yyyy-mm-dd"""
    m, d, y = us_date.split("/")
    return y + "-" + m + "-" + d


def sort_df_by_usdate(df: pd.DataFrame) -> pd.DataFrame:
    """Sort a DF by date - when the date has US format mm/dd/yyyy"""
    df["Sort_Date"] = df["Date"].apply(map_usdate_to_sortable_date)
    df.sort_values("Sort_Date", inplace=True, ascending=True)
    df.drop("Sort_Date", inplace=True, axis=1)
    df.reset_index(drop=True, inplace=True)
    return df


def write_pretty_venmo(in_df: pd.DataFrame, ven_file_name: str, sheet_name: str) -> None:
    """
    Writes to an Excel worksheet and formats nicely a DF whose values represent $
    """
    # Sort by Date
    tmp_df = in_df.copy(deep=True)
    df = sort_df_by_usdate(tmp_df)
    with pd.ExcelWriter(ven_file_name, mode="w", engine="openpyxl") as xlw:
        df.to_excel(xlw, sheet_name=sheet_name, float_format="%.2f", header=True, index=False)

    workbook = openpyxl.load_workbook(ven_file_name)
    sheet_idx = workbook.sheetnames.index(sheet_name)
    ws = workbook.worksheets[sheet_idx]
    dim_str = ws.dimensions
    col_names, max_row = expand_dimensions(dim_str)
    # Make 1st col and 1st row bold
    left_top_align = Alignment(horizontal="left", vertical="top")
    center_bottom_align = Alignment(horizontal="center", vertical="bottom")
    for row in range(1, max_row + 1):
        cel = "A" + str(row)
        ws[cel].alignment = left_top_align
    for col in col_names:
        cel = col + str(1)
        ws[cel].font = Font(bold=True)
        ws[cel].alignment = center_bottom_align

    idx = list(df.columns).index("Amount")
    amount_col = col_names[idx]
    for row_nb in range(2, max_row + 1):  # row 1 is the header
        cel = ws[amount_col + str(row_nb)]
        cel.number_format = '_("$"#,##0.00_);[Red]("$"#,##0.00);_($* "-"??_);_(@_)'

    # Reference: https://www.codespeedy.com/change-or-modify-column-width-size-in-openpyxl/
    # Reference: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.dimensions.html
    for col in col_names:
        ws.column_dimensions[col].bestFit = True

    # Freeze header row and column
    ws.freeze_panes = ws["B2"]
    ws.sheet_view.zoomScale = ZOOM_LEVEL
    workbook.save(ven_file_name)
    return None


def check_bad_tags(row: pd.Series, tag_lst: [str]) -> bool:
    """
    If Category starts with Travel - then Tag has to exist and be in the tag_lst
    :param row: [Category, Tag]
    :param tag_lst: list of tags
    :return: True if there is an error with the tag
    """
    cat, tag = list(row)
    if not cat:
        return True
    if str(cat).startswith("Travel"):
        if not tag:
            return True
        if tag not in tag_lst:
            return True
    return False


def check_bad_categories(vmo_file: str, bp_map_file: str) -> pd.DataFrame:
    """
    Check for bad catergies or tags in DF read from file vmo_file - based on data in bp_map_file
    :param vmo_file:
    :param bp_map_file:
    :return:
    """
    df = pd.read_excel(vmo_file, sheet_name=VENMO_STATEMENT_SHEET, header=0)
    # print(f"Venmo DF w/ NO flags:\n{df}")
    category_df = pd.read_excel(bp_map_file, sheet_name="Categories", header=0)
    category_lst = list(category_df["Category"])
    tags_df = pd.read_excel(bp_map_file, sheet_name="Travel", header=0)
    tags_lst = list(tags_df["Vacation"])
    # covers case where category is not filled
    df["bad_cat"] = df["Category"].apply(lambda x: x not in category_lst)
    df["bad_tags"] = df[["Category", "Tags"]].apply(lambda x: check_bad_tags(x, tags_lst), axis=1)
    # row is bad if cat or tag is bad
    df["bad"] = df[["bad_cat", "bad_tags"]].apply(lambda x: x[0] or x[1], axis=1)
    # print(f"Venmo DF w/ flags:\n{df}")
    bad_cat_df = df[df["bad"] == True].copy(deep=True)
    df.drop(["bad_cat", "bad_tags", "bad"], inplace=True, axis=1)
    return bad_cat_df


def find_most_recent_date(date_ser: pd.Series) -> str:
    """Find the most recent date in a Series of dates of format mm/dd/yyyy"""
    in_list = list(date_ser)
    tmp_list = [x.split("/") for x in in_list]  # list of [mm, dd, yyyy]
    out_list = [x[2] + "-" + x[0] + "-" + x[1] for x in tmp_list]  # yyyy-mm-dd
    max_date = max(out_list)
    y, m, d = max_date.split("-")
    return m + "/" + d + "/" + y  # mm/dd/yyyy


def read_venmo_file(in_file, venmo_debug: bool = False):
    """read Venmo account transactions from CSV file, clean up, and return in DF"""
    df = pd.read_csv(in_file, sep=",", header=VENMO_HEADER, index_col=None, usecols=VenmoCol2Keep)
    df.dropna(how="all", inplace=True, axis=0)
    if venmo_debug:
        if df.empty:
            print(f"File: {in_file} has NO NEW entries")
        else:
            print(f"File: {in_file} - DF:\n{df}")
    return df


def remap_venmo_date(in_datetime: str) -> str:
    """Remap from 2018-03-21T00:00:00 to 03/21/2018"""
    # There is not 'T' - dates may already be in the right format
    if len(in_datetime.split("T")) == 1:
        if len(in_datetime.split("/")) == 3:
            return in_datetime  # already formatted correctly
        # else:
        print(f"Error: Unexpected date format in Venmo file: {in_datetime}")
        sys.exit(3)
    in_date, _ = in_datetime.split("T")
    yy, mm, dd = in_date.split("-")
    return mm + "/" + dd + "/" + yy


def clean_up_venmo_df(in_df: pd.DataFrame) -> pd.DataFrame:
    """
    :param in_df:
    :return:
    """
    vmo_df = in_df.copy(deep=True)
    vmo_df["Datetime"] = vmo_df["Datetime"].map(remap_venmo_date)
    vmo_df["Payee"] = vmo_df[["Type", "From", "To"]].apply(set_payee, axis=1)
    vmo_df["Amount"] = vmo_df["Amount (total)"].apply(set_amount)
    vmo_df.drop(["Type", "From", "To", "Amount (total)"], inplace=True, axis=1)
    col_names = list(vmo_df.columns)
    new_col_names = [MapDict[x] for x in col_names]
    vmo_df.columns = new_col_names
    vmo_df["Tags"] = ""  # Add the Tags column
    vmo_df["Category"] = ""  # Add the Category column
    vmo_df["Source"] = "Venmo"  # Add the Source column
    vmo_df = vmo_df[BenPlanCol]  # Reorder the columns
    return vmo_df


def process_all_venmo_files(venmo_dir: str, venmo_xl_file: str, bp_map_file: str, skip_recon: bool = False,
                            venmo_debug: bool = False, ) -> pd.DataFrame:
    """
    :param venmo_dir:
    :param bp_map_file: file holding categories and tags
    :param skip_recon: True if the user wants to skip the reconciliation step
    :param venmo_debug: True if the user wants to see debug messages
    :return: file of all Venmo transactions to date
    """
    files_lst = [f.name for f in os.scandir(venmo_dir) if not f.is_dir()]
    # print(f"files_lst: {files_lst}")
    # fullmatch returns None if there is No match
    # Make a list of all the files that match the VENMO_NAME_PATTERN
    venmo_file_lst = [f for f in files_lst if re.fullmatch(VENMO_NAME_PATTERN, f) is not None]
    if venmo_debug:
        print(f"List of all Venmo files:\n{venmo_file_lst}")

    # Get the data of the latest entry in the already-processed file
    if os.path.exists(venmo_xl_file):
        venmo_todate_df = pd.read_excel(venmo_xl_file, sheet_name=VENMO_STATEMENT_SHEET, header=0)
        if venmo_debug:
            print(f"Read {len(venmo_todate_df.index)} Venmo Entries")
    else:
        venmo_todate_df = pd.DataFrame(columns=BenPlanCol)
   
    # Eliminate the files that have already been processed
    if len(venmo_todate_df.index) > 0:  # file not empty, let's get the date of most recent entry
        most_recent_date = find_most_recent_date(venmo_todate_df["Date"])
        print(f"most_recent_date in {venmo_xl_file}: {most_recent_date}")
        m, _, y = most_recent_date.split("/")
        most_recent_month = y + "_" + m
        files_to_process = make_list_of_files_to_process(venmo_file_lst, most_recent_month)
        print(f"Date of Most Recent Entry: {most_recent_date} - Most Recent Month: {most_recent_month}")
    else:  # No data aggregated so far
        files_to_process = venmo_file_lst

    if len(files_to_process) > 0:
        print(f"List of NEW Venmo files to process - count = {len(files_to_process)}:\n{files_to_process}")
    else:
        print("No NEW Venmo files to process")
        return venmo_todate_df
        
    if venmo_debug:
        print(f"{len(venmo_todate_df.index)} Existing Venmo Entries")
        print(f"List of all Venmo files - count = {len(venmo_file_lst)}:\n{venmo_file_lst}")
        print(f"\nList of files NOT PROCESSED - \
            count = {len(venmo_file_lst) - len(files_to_process)}:\n \
            {[f for f in venmo_file_lst if f not in files_to_process]}")


    # Process the new files
    vmo_cumul_df = pd.DataFrame(columns=VenmoCol2Keep)
    print(f"processing {len(files_to_process)} Venmno files: {files_to_process}")
    files_to_process = [venmo_dir + "/" + f for f in files_to_process]  # full pathname
    for venmo_file in files_to_process:
        one_file_df = read_venmo_file(venmo_file)
        if not one_file_df.empty:
            vmo_cumul_df = pd.concat([vmo_cumul_df, one_file_df], axis=0, ignore_index=True)

    if vmo_cumul_df.empty:  # there are NO new Venmo entries in the new files
        if venmo_debug:
            print("No NEW Entries to process in the new files")
        return venmo_todate_df

    # Simplify Datetime, Figure out payee and amount for the new entries
    new_venmo_df = clean_up_venmo_df(vmo_cumul_df)
    if venmo_debug:
        print(f"{len(new_venmo_df.index)} entries to Categorize & Tag in Venmo file")
    # Combine entries to date with new ones
    vmo_df = pd.concat([venmo_todate_df, new_venmo_df], ignore_index=True, axis=0)
    # Overwrite Venmo file with new data
    write_pretty_venmo(vmo_df, venmo_xl_file, sheet_name=VENMO_STATEMENT_SHEET)

    # User has to enter the categories and tags offline
    if not skip_recon:  # Don't skip the reconciliation
        done_flag = False
        while not done_flag:
            print(f"\nEnter Categories and Tags offline in file {venmo_xl_file}")
            print(" Categorize as Transfer To Chase in the Venmo_statements.xlsx file - and -",
                  "... Venmo Cashout Ppd in Quicken")
            print(" Categorize as Venmo_Dupe in Quicken transactions \
                that are also the Venmo_statements.xlsx file ")
            print("Add Tag Dinna_Venmo for Venmo transactions to/from Dinna's account - \
                ... unless they can be attributed specifically \
                (e.g. travel expense reimbursement) ")
            # done_flag is True when input is 'yes'
            done_flag = input("Enter Yes when done ->  ").lower() == "yes"

    # Check if the Categories and Tags have been entered properly
    bad_categories_df = check_bad_categories(venmo_xl_file, bp_map_file)
    if len(bad_categories_df.index) > 0:
        print(f"\nERROR: FIX {len(bad_categories_df.index)} \
        transactions for bad categories or bad Tags")
        # need to make indices sequential for loop below
        bad_categories_df.reset_index(drop=True, inplace=True)
        for idx in bad_categories_df.index:  # make sure all rows are printed
            # print(bad_categories_df.iloc[idx])
            print(f"{show_values(bad_categories_df.columns, bad_categories_df.iloc[idx].values)}\n")
            sys.exit(2)
    else:
        print("Venmo Categories & Tags all check out")

    return vmo_df


if __name__ == "__main__":
    # execute only if run as a script

    home_dir = '/Users/bfraenkel/Documents/Code/BenPlan/'
    venmo_dir = home_dir + 'Data/Venmo'
    benplan_map_file = home_dir + 'BenPlan_Maps.xlsx'
    venmo_xl_file = venmo_dir + '/Venmo_statements.xlsx'
    venmo_df = process_all_venmo_files(venmo_dir, venmo_xl_file, benplan_map_file, venmo_debug=False)
    print(f"Result Venmo DF:\n{venmo_df}")

    sys.exit(0)
