import asyncio
import pandas as pd
from multiprocessing import Pipe
import openpyxl
from openpyxl.styles import Font, Alignment
import string
import time
import os
from utils import this_function, now_str
from logger import logger
from typing import Optional, Tuple, List
from tick_timer_class import TickTimer

ZOOM_LEVEL = 120
QUICK_DIR = 'Quick'
SKIP_WRITE_NICE = False  # use to debug the code -> nothing is written to excel

async def _write_nice_async(conn: Pipe, *args) -> None:
    """
    Async implementation of sending data through a pipe. The main process does not wait
    for the function to complete
    The pipe receives 5-tuple: xl_out_filename, df, sheet, index_flag, mv_last
    :return: None

    References:
    """

    before = time.perf_counter()
    conn.send(args)  # FYI: no '*'
    delta = time.perf_counter() - before
    # args[0] is the sheet name
    logger.info(f"{this_function()}: {args[2]} - {now_str()} - {delta:,.6f} - pid: {os.getpid()}")
    return


def double_write(conn: Pipe, xl_out_filename: str | None, df_in: pd.DataFrame,
                 sheetname: str, index_flag=False, mv_last=None, quick_flag=True) -> asyncio.Task:
    """
    Sends the DF to the async process to be written to an Excel file
    :param conn: Pipe to send the data to the async process
    :param xl_out_filename: 
    :param df_in: 
    :param sheetname: 
    :param index_flag: 
    :param mv_last: 
    :param quick_flag: 
    :return: asyncio.Task
    """
    # a None output file is a signal to close the connection / nothing to write
    if xl_out_filename is  None:
        logger.info(f"double_write: closing connection")
        task = asyncio.create_task(
            _write_nice_async(conn, None, None, None, None, None))
        return task
    
    if quick_flag: # write unformatted data as is to a quick file
        df_in.to_excel(f"{QUICK_DIR}/{sheetname}.xlsx", sheet_name=sheetname, float_format='%.2f',
                       header=True, index=index_flag)

    # set up async process to format and write the data
    task = asyncio.create_task(
            _write_nice_async(conn, xl_out_filename, df_in, sheetname, index_flag, mv_last))
    return task


def write_nice_df_2_xl_pipe(conn: Pipe) -> None:
    """
    Consumes the tasks received from the async process
    The pipe receives 5-tuple: xl_out_filename, df, sheet, index_flag, mv_last

    Close the pipe by sending None as filename
    Writes to an Excel worksheet and formats nicely a DF whose values represent $
    Reference: https://www.codespeedy.com/change-or-modify-column-width-size-in-openpyxl/
    Reference: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.dimensions.html
    """
    while True:  # loop until the pipe is closed
        xl_out_filename, df, sheet, index_flag, mv_last = conn.recv()
        if xl_out_filename is None:  # Done -> exit
            logger.info(f"{this_function()}: DONE - {now_str()} - pid: {os.getpid()}")
            break
        # Process the DF
        logger.info(f"{this_function()}: {sheet} - {now_str()} - pid: {os.getpid()}")
        _write_nice_df_2_xl(xl_out_filename, df, sheet, index_flag, mv_last)
        
    return

def _write_nice_df_2_xl(xl_out_filename: str, df: pd.DataFrame, sheet: str, index_flag: bool = False,
                        mv_last: Optional[int] = None) -> None:
    """Write DataFrame to Excel with nice formatting."""
    write_nice_timer = TickTimer(name="write_nice_df_2_xl")
    write_nice_timer.start()
    if mv_last is not None:  # Move last columns to front
            df = _mv_last_n_col_to_front(df, mv_last)
    mode_a_w = 'a' if os.path.exists(xl_out_filename) else 'w'
    # Write the data non-formatted
    with pd.ExcelWriter(xl_out_filename, mode=mode_a_w, engine='openpyxl') as xlw:
        df.to_excel(xlw, sheet_name=sheet, float_format='%.2f', header=True, index=index_flag,
                        freeze_panes=(1, 1))

    if SKIP_WRITE_NICE:  # Skip formatting
        write_nice_timer.tick(sheet)
        return

    # re-open the same workbook to format it
    workbook = openpyxl.load_workbook(xl_out_filename)
    sheet_idx = workbook.sheetnames.index(sheet)
    ws = workbook.worksheets[sheet_idx]
    dim_str = ws.dimensions
    col_names, max_row = expand_dimensions(dim_str)
    left_top_align = Alignment(horizontal='left', vertical='top')
    center_bottom_align = Alignment(horizontal='center', vertical='bottom')
    for row in range(1, max_row + 1):
        cel = 'A' + str(row)
        ws[cel].font = Font(bold=True)
        ws[cel].alignment = left_top_align
    for i, col in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True)
        cell.alignment = center_bottom_align

    for col in ws.iter_cols(min_row=2, max_row=max_row, min_col=2, max_col=len(col_names)):
        for cel in col:
            # ws[cel].number_format = 'Currency'
            cel.number_format = '_("$"#,##0.00_);[Red]("$"#,##0.00);_($* "-"??_);_(@_)'

    for col in col_names:
        ws.column_dimensions[col].bestFit = True

    # Freeze header row and column
    if mv_last:
        freeze_col = string.ascii_uppercase[2 + mv_last - 1]  # 0-indexing
    else:
        freeze_col = 'B'

    ws.freeze_panes = ws[freeze_col + str(2)]
    ws.sheet_view.zoomScale = ZOOM_LEVEL
    workbook.save(xl_out_filename)
    write_nice_timer.tick(sheet)
    return

def expand_dimensions(dim_str: str) -> Tuple[List[str], int]:
    """Expand Excel dimension string to column names and row count."""
    t_l, b_r = dim_str.split(':')
    assert t_l == "A1", f"Top left {t_l} is not A1"

    last_col = ''
    digits = []
    for c in b_r:
        if c.isalpha():
            last_col += c
        elif c.isnumeric():
            digits.append(c)

    digits.reverse()
    powr_10 = 1
    last_row = 0
    for d in digits:
        last_row += int(d) * powr_10
        powr_10 *= 10

    col_range = []
    for ll in string.ascii_uppercase:
        col_range.append(ll)
        if ll == last_col:
            break

    if len(last_col) == 2:
        for l1 in string.ascii_uppercase:
            for l2 in string.ascii_uppercase:
                ll = l1 + l2
                col_range.append(ll)
                if ll == last_col:
                    break
            if ll == last_col:
                break

    return col_range, last_row

def _mv_last_n_col_to_front(df: pd.DataFrame, n: int) -> pd.DataFrame:
    """Move last n columns to front."""
    col = list(df.columns)
    new_cols = col[len(col) - n:] + col[:-n]
    result = df[new_cols]
    assert isinstance(result, pd.DataFrame)
    return result

