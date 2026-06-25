#!/usr/local/bin/python3
"""
Extracts JPMorgan Capital Market Projections 2026 data into two worksheets:
    - Stats: Asset Type, Asset Class, Expected Return, Standard Deviation (rows 9-67)
    - Correlation: Asset Class labels + correlation matrix (rows 9-67, cols G-BM)

FIXME: the file paths are hard-coded for now, but should be configurable in the future.
"""

import openpyxl
import pandas as pd
import sys
import os

DATA_DIRECTORY = "/Users/bfraenkel/Documents/Code/Financial_Plan/Data/JPM"
SOURCE_FILE = os.path.join(DATA_DIRECTORY, "JPMorgan Capital Market Projections 2026.xlsx")
OUTPUT_FILE = os.path.join(DATA_DIRECTORY, "jpm_cpm_stats_2026_06_16.xlsx")

DATA_ROWS = range(9, 68)       # rows 9-67 inclusive
CORR_COL_START = 7             # column G
CORR_COL_END = 65              # column BM
CORR_HEADER_ROW = 5            # asset class names for correlation columns are in row 5


def extract(src_path: str, out_path: str) -> None:
    src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = src.active
    if ws_src is None:
        raise ValueError("Unable to load active worksheet from source file")

    asset_classes = [ws_src.cell(r, 2).value for r in DATA_ROWS]

    stats_df = pd.DataFrame(
        {
            "Asset Class": asset_classes,   # Important that this is the first column so that it becomes the index when read back
            # "Asset Type": [ws_src.cell(r, 1).value for r in DATA_ROWS],
            "Expected Return": [ws_src.cell(r, 3).value for r in DATA_ROWS],
            "Standard Deviation": [ws_src.cell(r, 5).value for r in DATA_ROWS],
        }
    )
    stats_df["Asset Type"] = stats_df["Asset Type"].ffill()

    corr_df = pd.DataFrame(
        [
            [ws_src.cell(r, c).value for c in range(CORR_COL_START, CORR_COL_END + 1)]
            for r in DATA_ROWS
        ],
        index=asset_classes,
        columns=asset_classes,
    )
    # Fill in the top half of the correlation matrix with the lower half 
    for i in range(len(asset_classes)):
        for j in range(i + 1, len(asset_classes)):
            corr_df.iloc[i, j] = corr_df.iloc[j, i]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        stats_df.to_excel(writer, sheet_name="Stats", index=False)
        corr_df.to_excel(writer, sheet_name="Correlation")

    print(f"Saved: {out_path}")


def main(cmd_line: list[str]) -> None:
    extract(SOURCE_FILE, OUTPUT_FILE)


if __name__ == "__main__":
    main(sys.argv)
    sys.exit("---\nDone!")
